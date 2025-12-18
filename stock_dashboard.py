import argparse
import logging
import os
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import yfinance as yf

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as e:
    raise SystemExit(
        "Missing dependency: openpyxl. Install with: pip install openpyxl"
    ) from e


LOGGER = logging.getLogger("stock_dashboard")


@dataclass(frozen=True)
class IndicatorConfig:
    rsi_period: int = 14
    atr_period: int = 14
    sma_fast: int = 20
    sma_slow: int = 50


def _setup_logging(log_level: str) -> None:
    level = getattr(logging, log_level.upper(), None)
    if level is None:
        raise ValueError(f"Invalid log level: {log_level}")
    logging.basicConfig(
        level=level,
        format="%(asctime)s %(levelname)s %(name)s - %(message)s",
    )


def _parse_tickers(raw: str) -> List[str]:
    tickers = [t.strip() for t in raw.split(",") if t.strip()]
    if not tickers:
        raise ValueError("No tickers provided. Example: --tickers AAPL,MSFT,RELIANCE.NS")
    if len(tickers) > 200:
        raise ValueError("Refusing to run with >200 tickers in one go (safety limit).")
    return tickers


def fetch_ohlcv(
    ticker: str,
    start: str,
    end: str,
    auto_adjust: bool = True,
) -> pd.DataFrame:
    """
    Fetch daily OHLCV from yfinance. Fails loudly if empty.
    """
    LOGGER.info("Fetching %s (%s → %s, auto_adjust=%s)", ticker, start, end, auto_adjust)
    df = yf.Ticker(ticker).history(start=start, end=end, auto_adjust=auto_adjust, interval="1d")
    if df is None or df.empty:
        raise ValueError(f"No data returned for ticker: {ticker}")
    required = {"Open", "High", "Low", "Close", "Volume"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Ticker {ticker}: missing columns {sorted(missing)}")
    df = df.copy()
    df.index = pd.to_datetime(df.index)
    df.sort_index(inplace=True)
    return df


def compute_indicators(df: pd.DataFrame, cfg: IndicatorConfig) -> pd.DataFrame:
    """
    Adds: daily_return, log_return, RSI, ATR, MACD, signal, SMA/EMA, volume_sma, volume_ratio
    """
    out = df.copy()

    # Returns
    out["daily_return"] = out["Close"].pct_change(fill_method=None)
    out["log_return"] = np.log(out["Close"] / out["Close"].shift(1))

    # RSI (Wilder-ish using EMA smoothing)
    delta = out["Close"].diff()
    gain = delta.where(delta > 0, 0.0)
    loss = (-delta.where(delta < 0, 0.0))

    avg_gain = gain.ewm(alpha=1 / cfg.rsi_period, adjust=False).mean()
    avg_loss = loss.ewm(alpha=1 / cfg.rsi_period, adjust=False).mean().replace(0, np.nan)
    rs = avg_gain / avg_loss
    out["rsi"] = 100 - (100 / (1 + rs))
    out["rsi"] = out["rsi"].fillna(0.0).clip(lower=0.0, upper=100.0)

    # ATR
    hl = out["High"] - out["Low"]
    hc = (out["High"] - out["Close"].shift(1)).abs()
    lc = (out["Low"] - out["Close"].shift(1)).abs()
    tr = pd.concat([hl, hc, lc], axis=1).max(axis=1)
    out["atr"] = tr.ewm(alpha=1 / cfg.atr_period, adjust=False).mean()

    # MACD
    ema12 = out["Close"].ewm(span=12, adjust=False).mean()
    ema26 = out["Close"].ewm(span=26, adjust=False).mean()
    out["macd"] = ema12 - ema26
    out["macd_signal"] = out["macd"].ewm(span=9, adjust=False).mean()
    out["macd_hist"] = out["macd"] - out["macd_signal"]

    # SMAs/EMAs
    out[f"sma_{cfg.sma_fast}"] = out["Close"].rolling(cfg.sma_fast).mean()
    out[f"sma_{cfg.sma_slow}"] = out["Close"].rolling(cfg.sma_slow).mean()
    out[f"ema_{cfg.sma_fast}"] = out["Close"].ewm(span=cfg.sma_fast, adjust=False).mean()

    # Volume ratio
    out["volume_sma_20"] = out["Volume"].rolling(20).mean()
    out["volume_ratio"] = (out["Volume"] / out["volume_sma_20"]).replace([np.inf, -np.inf], np.nan).fillna(0.0)

    return out


def build_latest_snapshot(
    df: pd.DataFrame,
    ticker: str,
    cfg: IndicatorConfig,
) -> Dict[str, object]:
    if df.empty:
        raise ValueError(f"{ticker}: empty dataframe at snapshot stage")

    last = df.iloc[-1]
    close = float(last["Close"])
    rsi = float(last.get("rsi", np.nan))
    atr = float(last.get("atr", np.nan))
    vol_ratio = float(last.get("volume_ratio", np.nan))

    def _ret(n: int) -> float:
        if len(df) <= n:
            return float("nan")
        prev = df["Close"].iloc[-(n + 1)]
        return float((close / prev) - 1.0)

    snapshot = {
        "ticker": ticker,
        "asof": df.index[-1].strftime("%Y-%m-%d"),
        "close": close,
        "ret_1d": _ret(1),
        "ret_5d": _ret(5),
        "ret_20d": _ret(20),
        "rsi": rsi,
        "atr": atr,
        "vol_ratio": vol_ratio,
        f"sma_{cfg.sma_fast}": float(last.get(f"sma_{cfg.sma_fast}", np.nan)),
        f"sma_{cfg.sma_slow}": float(last.get(f"sma_{cfg.sma_slow}", np.nan)),
        "macd": float(last.get("macd", np.nan)),
        "macd_signal": float(last.get("macd_signal", np.nan)),
    }
    return snapshot


def _autosize_columns(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(v))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def write_excel_dashboard(
    out_path: str,
    snapshots: List[Dict[str, object]],
    metadata: Dict[str, str],
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dashboard"

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    ws["A1"] = "Stock Dashboard"
    ws["A1"].font = title_font
    ws["A2"] = f"Generated at: {metadata['generated_at']}"
    ws["A3"] = f"Date range: {metadata['start']} → {metadata['end']}"
    ws["A4"] = f"Tickers: {metadata['tickers']}"

    start_row = 6
    cols = [
        "ticker", "asof", "close",
        "ret_1d", "ret_5d", "ret_20d",
        "rsi", "atr", "vol_ratio",
        "macd", "macd_signal",
    ]

    for j, c in enumerate(cols, 1):
        cell = ws.cell(start_row, j, c)
        cell.font = header_font
        cell.fill = header_fill

    for i, snap in enumerate(snapshots, 1):
        r = start_row + i
        for j, c in enumerate(cols, 1):
            ws.cell(r, j, snap.get(c))

    _autosize_columns(ws)

    # Optional: write a raw sheet too (helps show “real engineering”)
    ws2 = wb.create_sheet("Notes")
    ws2["A1"] = "This dashboard is a generic, educational sample."
    ws2["A2"] = "It demonstrates data ingestion, indicator computation, and reporting."
    ws2["A3"] = "It is NOT financial advice."

    wb.save(out_path)
    LOGGER.info("Wrote Excel dashboard: %s", out_path)


def ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def main() -> int:
    parser = argparse.ArgumentParser(description="Generic Stock Dashboard Generator (CSV + Excel)")
    parser.add_argument("--tickers", required=True, help="Comma-separated tickers, e.g. AAPL,MSFT,RELIANCE.NS")
    parser.add_argument("--start", required=True, help="Start date YYYY-MM-DD")
    parser.add_argument("--end", required=True, help="End date YYYY-MM-DD (exclusive in yfinance history)")
    parser.add_argument("--out-dir", default="output", help="Output directory")
    parser.add_argument("--auto-adjust", action="store_true", help="Use adjusted prices (recommended)")
    parser.add_argument("--log-level", default="INFO", help="DEBUG/INFO/WARNING/ERROR")
    args = parser.parse_args()

    _setup_logging(args.log_level)

    # Validate dates early (fail loudly)
    try:
        datetime.strptime(args.start, "%Y-%m-%d")
        datetime.strptime(args.end, "%Y-%m-%d")
    except ValueError as e:
        raise SystemExit("Invalid date format. Use YYYY-MM-DD.") from e

    tickers = _parse_tickers(args.tickers)
    ensure_dir(args.out_dir)

    cfg = IndicatorConfig()
    snapshots: List[Dict[str, object]] = []

    combined_rows: List[pd.DataFrame] = []
    errors: List[Tuple[str, str]] = []

    for t in tickers:
        try:
            df = fetch_ohlcv(t, args.start, args.end, auto_adjust=args.auto_adjust)
            df_ind = compute_indicators(df, cfg)
            df_ind["ticker"] = t
            df_ind.reset_index(inplace=True)
            df_ind.rename(columns={"index": "Date"}, inplace=True)

            # CSV outputs
            per_ticker_csv = os.path.join(args.out_dir, f"{t.replace('/', '_')}_data.csv")
            df_ind.to_csv(per_ticker_csv, index=False)

            snapshots.append(build_latest_snapshot(df_ind.set_index("Date"), t, cfg))
            combined_rows.append(df_ind)

        except Exception as e:
            msg = str(e)
            LOGGER.error("Failed ticker=%s error=%s", t, msg)
            errors.append((t, msg))

    if not combined_rows:
        LOGGER.error("No tickers succeeded. Exiting.")
        for t, msg in errors:
            LOGGER.error("ticker=%s error=%s", t, msg)
        return 2

    combined = pd.concat(combined_rows, ignore_index=True)
    combined_csv = os.path.join(args.out_dir, "combined_data.csv")
    combined.to_csv(combined_csv, index=False)
    LOGGER.info("Wrote combined CSV: %s", combined_csv)

    snapshots_df = pd.DataFrame(snapshots).sort_values(["ticker"])
    snapshots_csv = os.path.join(args.out_dir, "latest_snapshot.csv")
    snapshots_df.to_csv(snapshots_csv, index=False)
    LOGGER.info("Wrote snapshot CSV: %s", snapshots_csv)

    # Excel dashboard
    dashboard_xlsx = os.path.join(args.out_dir, "stock_dashboard.xlsx")
    write_excel_dashboard(
        dashboard_xlsx,
        snapshots=snapshots,
        metadata={
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "start": args.start,
            "end": args.end,
            "tickers": ",".join(tickers),
        },
    )

    # Persist errors for transparency (also looks good professionally)
    if errors:
        err_path = os.path.join(args.out_dir, "errors.csv")
        pd.DataFrame(errors, columns=["ticker", "error"]).to_csv(err_path, index=False)
        LOGGER.warning("Some tickers failed; wrote errors to %s", err_path)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
